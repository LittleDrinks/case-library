#!/usr/bin/env python3
"""Back-office account administration commands."""

# ruff: noqa: E402

import argparse
import csv
import os
import secrets
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

BACKEND_DIR = Path(__file__).resolve().parents[1]
REPO_ROOT = BACKEND_DIR.parent
sys.path.insert(0, str(REPO_ROOT))
sys.path.insert(0, str(BACKEND_DIR))

from backend.db.connection import init_db
from backend.db.constants import USER_ROLES, USER_STATUSES
from backend.repositories.users import (
    clear_users,
    create_user,
    delete_user,
    get_user_by_username,
    list_users,
    rename_user,
    set_user_password,
    update_user_fields,
)

IDENTITY_COLUMNS = ("username", "school_id")
DISPLAY_NAME_COLUMNS = ("nickname", "display_name", "name")
SUPPORTED_IMPORT_COLUMNS = {
    "username",
    "school_id",
    "password",
    "role",
    "nickname",
    "display_name",
    "name",
    "must_change_password",
    "status",
    "department",
    "class",
    "organization",
}
UNSUPPORTED_METADATA_COLUMNS = ("department", "class", "organization")
DEFAULT_IMPORT_STATUS = "active"
DEFAULT_IMPORT_MUST_CHANGE_PASSWORD = True
MIN_IMPORT_PASSWORD_LENGTH = 8


@dataclass
class ImportRow:
    row_number: int
    username: str
    password: str
    role: str
    nickname: str
    must_change_password: bool
    status: str
    password_generated: bool = False
    ignored_columns: list[str] = field(default_factory=list)
    source_identity: str = "username"


@dataclass
class ImportSummary:
    total_rows: int = 0
    valid_rows: int = 0
    would_create: int = 0
    created: int = 0
    skipped_existing: int = 0
    errors: int = 0
    generated_passwords: int = 0

    def as_dict(self) -> dict[str, int]:
        return {
            "total_rows": self.total_rows,
            "valid_rows": self.valid_rows,
            "would_create": self.would_create,
            "created": self.created,
            "skipped_existing": self.skipped_existing,
            "errors": self.errors,
            "generated_passwords": self.generated_passwords,
        }


def parse_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in ("1", "true", "yes", "y", "是"):
        return True
    if normalized in ("0", "false", "no", "n", "否"):
        return False
    raise argparse.ArgumentTypeError(f"Invalid boolean value: {value}")


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(value: Any) -> str:
    return normalize_cell(value).lower().replace("-", "_").replace(" ", "_")


def normalize_role(value: str) -> str:
    role = normalize_cell(value).lower()
    if role == "user":
        role = "normal"
    return role


def read_csv_rows(path: Path, encoding: str) -> tuple[list[dict[str, str]], list[str]]:
    with path.open(encoding=encoding, newline="") as file:
        reader = csv.DictReader(file)
        fieldnames = [normalize_header(name) for name in (reader.fieldnames or [])]
        rows = []
        for row in reader:
            rows.append(
                {
                    normalize_header(key): normalize_cell(value)
                    for key, value in row.items()
                    if key is not None
                }
            )
        return rows, fieldnames


def read_xlsx_rows(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            "XLSX import requires openpyxl. Install runtime dependencies from requirements.txt."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        workbook.close()
        return [], []

    fieldnames = [normalize_header(cell) for cell in header_row]
    rows = []
    for values in rows_iter:
        row = {}
        for index, header in enumerate(fieldnames):
            if not header:
                continue
            value = values[index] if index < len(values) else None
            row[header] = normalize_cell(value)
        rows.append(row)
    workbook.close()
    return rows, fieldnames


def read_import_rows(path: Path, encoding: str) -> tuple[list[dict[str, str]], list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(path, encoding)
    if suffix == ".xlsx":
        return read_xlsx_rows(path)
    raise SystemExit("Unsupported import file type. Use .csv or .xlsx.")


def pick_first(row: dict[str, str], columns: tuple[str, ...]) -> tuple[str, str]:
    for column in columns:
        value = normalize_cell(row.get(column))
        if value:
            return value, column
    return "", columns[0]


def generate_import_password() -> str:
    return secrets.token_urlsafe(18)


def parse_import_row(
    row_number: int,
    row: dict[str, str],
    default_password: str | None,
    missing_password: str,
) -> tuple[ImportRow | None, list[str]]:
    errors = []
    username, source_identity = pick_first(row, IDENTITY_COLUMNS)
    nickname, _ = pick_first(row, DISPLAY_NAME_COLUMNS)
    role = normalize_role(row.get("role", ""))
    status = normalize_cell(row.get("status")).lower() or DEFAULT_IMPORT_STATUS
    password = normalize_cell(row.get("password"))
    password_generated = False

    if not username:
        errors.append("missing username or school_id")
    if not nickname:
        errors.append("missing nickname, display_name, or name")
    if not role:
        errors.append("missing role")
    elif role not in USER_ROLES:
        errors.append(f"invalid role: {role}")
    if status not in USER_STATUSES:
        errors.append(f"invalid status: {status}")

    must_change_password_value = row.get("must_change_password", "")
    try:
        must_change_password = (
            parse_bool(must_change_password_value)
            if must_change_password_value != ""  # nosec B105
            else DEFAULT_IMPORT_MUST_CHANGE_PASSWORD
        )
    except argparse.ArgumentTypeError as exc:
        must_change_password = DEFAULT_IMPORT_MUST_CHANGE_PASSWORD
        errors.append(str(exc))

    if not password and default_password:
        password = default_password
    if not password:
        if missing_password == "error":  # nosec B105
            errors.append("missing password")
        else:
            password = generate_import_password()
            password_generated = True

    if password and len(password) < MIN_IMPORT_PASSWORD_LENGTH:
        errors.append(f"password must be at least {MIN_IMPORT_PASSWORD_LENGTH} characters")

    ignored_columns = [
        column for column in UNSUPPORTED_METADATA_COLUMNS if normalize_cell(row.get(column))
    ]

    if errors:
        return None, errors
    return (
        ImportRow(
            row_number=row_number,
            username=username,
            password=password,
            role=role,
            nickname=nickname,
            must_change_password=must_change_password,
            status=status,
            password_generated=password_generated,
            ignored_columns=ignored_columns,
            source_identity=source_identity,
        ),
        [],
    )


def validate_import_file_headers(fieldnames: list[str]) -> list[str]:
    errors = []
    fieldname_set = set(fieldnames)
    if not fieldname_set.intersection(IDENTITY_COLUMNS):
        errors.append("missing username or school_id column")
    if not fieldname_set.intersection(DISPLAY_NAME_COLUMNS):
        errors.append("missing nickname, display_name, or name column")
    if "role" not in fieldname_set:
        errors.append("missing role column")
    unknown_columns = sorted(fieldname_set - SUPPORTED_IMPORT_COLUMNS - {""})
    if unknown_columns:
        errors.append(f"unsupported columns: {', '.join(unknown_columns)}")
    return errors


def build_import_plan(
    raw_rows: list[dict[str, str]],
    default_password: str | None,
    missing_password: str,
) -> tuple[list[ImportRow], list[str], ImportSummary]:
    parsed_rows = []
    errors = []
    summary = ImportSummary(total_rows=len(raw_rows))
    seen_usernames: dict[str, int] = {}

    for offset, row in enumerate(raw_rows, start=2):
        parsed, row_errors = parse_import_row(
            row_number=offset,
            row=row,
            default_password=default_password,
            missing_password=missing_password,
        )
        if row_errors:
            summary.errors += 1
            errors.append(f"Row {offset}: {'; '.join(row_errors)}")
            continue

        if parsed is None:
            summary.errors += 1
            errors.append(f"Row {offset}: import row could not be parsed")
            continue
        first_seen = seen_usernames.get(parsed.username)
        if first_seen is not None:
            summary.errors += 1
            errors.append(
                f"Row {offset}: duplicate username in import file: "
                f"{parsed.username} (first seen on row {first_seen})"
            )
            continue
        seen_usernames[parsed.username] = offset
        parsed_rows.append(parsed)

    summary.valid_rows = len(parsed_rows)
    summary.generated_passwords = sum(1 for row in parsed_rows if row.password_generated)
    return parsed_rows, errors, summary


def print_import_summary(summary: ImportSummary):
    print(
        "Import summary: "
        f"total_rows={summary.total_rows}, "
        f"valid_rows={summary.valid_rows}, "
        f"would_create={summary.would_create}, "
        f"created={summary.created}, "
        f"skipped_existing={summary.skipped_existing}, "
        f"errors={summary.errors}, "
        f"generated_passwords={summary.generated_passwords}"
    )


def print_users():
    users = list_users()
    if not users:
        print("No accounts found.")
        return
    for user in users:
        print(
            f"{user.get('username')}\t"
            f"{user.get('role')}\t"
            f"{user.get('nickname', '')}\t"
            f"must_change_password={user.get('must_change_password')}\t"
            f"status={user.get('status')}"
        )


def create_account(args):
    user_id = create_user(
        username=args.username,
        password=args.password,
        role=args.role,
        nickname=args.nickname,
        must_change_password=args.must_change_password,
        status=args.status,
    )
    print(f"Created account {args.username} with id={user_id}")


def import_users(args):
    path = Path(args.file)
    raw_rows, fieldnames = read_import_rows(path, args.encoding)
    header_errors = validate_import_file_headers(fieldnames)
    if header_errors:
        for error in header_errors:
            print(f"Import error: {error}")
        raise SystemExit(2)

    default_password = None
    if args.default_password_env:
        default_password = normalize_cell(os.environ.get(args.default_password_env))
        if not default_password:
            raise SystemExit(f"Environment variable is empty: {args.default_password_env}")
        if len(default_password) < MIN_IMPORT_PASSWORD_LENGTH:
            raise SystemExit(
                f"Default password must be at least {MIN_IMPORT_PASSWORD_LENGTH} characters"
            )

    parsed_rows, row_errors, summary = build_import_plan(
        raw_rows=raw_rows,
        default_password=default_password,
        missing_password=args.missing_password,
    )
    for error in row_errors:
        print(f"Import error: {error}")

    ignored_columns = sorted({column for row in parsed_rows for column in row.ignored_columns})
    if ignored_columns:
        print(
            "Import warning: current users schema does not store columns: "
            f"{', '.join(ignored_columns)}"
        )

    has_validation_errors = summary.errors > 0
    for row in parsed_rows:
        if get_user_by_username(row.username):
            summary.skipped_existing += 1
            print(f"Skip existing account: {row.username}")
            continue

        summary.would_create += 1
        if args.dry_run or has_validation_errors:
            print(f"Would create account: {row.username}")
            continue

        try:
            create_user(
                username=row.username,
                password=row.password,
                role=row.role,
                nickname=row.nickname,
                must_change_password=row.must_change_password,
                status=row.status,
            )
            summary.created += 1
            print(f"Created account: {row.username}")
        except Exception as exc:
            summary.errors += 1
            print(f"Import error: Row {row.row_number}: failed to create {row.username}: {exc}")

    if summary.generated_passwords:
        if args.dry_run:
            print(
                "Import warning: temporary passwords would be generated but are not printed."
            )
        else:
            print(
                "Import warning: generated temporary passwords were stored but not printed. "
                "Reset those accounts before distributing credentials."
            )

    print_import_summary(summary)
    if summary.errors:
        raise SystemExit(1)


def reset_password(args):
    if not set_user_password(args.username, args.password, must_change_password=True):
        raise SystemExit(f"Account not found: {args.username}")
    print(f"Password reset for {args.username}. must_change_password=true")


def delete_account(args):
    if not delete_user(args.username):
        raise SystemExit(f"Account not found: {args.username}")
    print(f"Deleted account: {args.username}")


def clear_accounts(args):
    if not args.yes:
        raise SystemExit("Refusing to clear accounts without --yes")
    deleted = clear_users()
    print(f"Deleted {deleted} accounts.")


def update_account(args):
    if not update_user_fields(
        args.username, role=args.role, nickname=args.nickname, status=args.status
    ):
        raise SystemExit(f"Account not found or no changes: {args.username}")
    print(f"Updated account: {args.username}")


def rename_account(args):
    if not rename_user(args.old_username, args.new_username):
        raise SystemExit(f"Account not found or no changes: {args.old_username}")
    print(f"Renamed account {args.old_username} -> {args.new_username}")


def build_parser():
    parser = argparse.ArgumentParser(description="Manage case library accounts")
    sub = parser.add_subparsers(dest="command", required=True)

    sub.add_parser("list", help="List all accounts").set_defaults(func=lambda args: print_users())

    create = sub.add_parser("create", help="Create one account")
    create.add_argument("--username", required=True)
    create.add_argument("--password", required=True)
    create.add_argument("--role", choices=["normal", "admin"], required=True)
    create.add_argument("--nickname", default="")
    create.add_argument("--must-change-password", type=parse_bool, default=True)
    create.add_argument("--status", choices=["active", "no_active"], default="active")
    create.set_defaults(func=create_account)

    import_parent = argparse.ArgumentParser(add_help=False)
    import_parent.add_argument("--file", required=True)
    import_parent.add_argument("--encoding", default="utf-8-sig")
    import_parent.add_argument("--dry-run", action="store_true")
    import_parent.add_argument(
        "--missing-password",
        choices=["generate", "error"],
        default="generate",
        help="Generate an undisclosed temporary password or reject rows without passwords.",
    )
    import_parent.add_argument(
        "--default-password-env",
        help="Read a fallback password from this environment variable without printing it.",
    )

    bulk = sub.add_parser(
        "import-users",
        parents=[import_parent],
        help="Import accounts from CSV or XLSX without opening public registration",
    )
    bulk.set_defaults(func=import_users)

    bulk_csv = sub.add_parser(
        "import-csv",
        parents=[import_parent],
        help="Compatibility alias for import-users; accepts .csv or .xlsx files",
    )
    bulk_csv.set_defaults(func=import_users)

    reset = sub.add_parser("reset-password", help="Reset one account password")
    reset.add_argument("--username", required=True)
    reset.add_argument("--password", required=True)
    reset.set_defaults(func=reset_password)

    delete = sub.add_parser("delete", help="Delete one account")
    delete.add_argument("--username", required=True)
    delete.set_defaults(func=delete_account)

    clear = sub.add_parser("clear", help="Delete all accounts")
    clear.add_argument("--yes", action="store_true")
    clear.set_defaults(func=clear_accounts)

    update = sub.add_parser("update", help="Update role, nickname, or status")
    update.add_argument("--username", required=True)
    update.add_argument("--role", choices=["normal", "admin"])
    update.add_argument("--nickname")
    update.add_argument("--status", choices=["active", "no_active"])
    update.set_defaults(func=update_account)

    rename = sub.add_parser("rename", help="Rename username")
    rename.add_argument("--old-username", required=True)
    rename.add_argument("--new-username", required=True)
    rename.set_defaults(func=rename_account)

    return parser


def main():
    parser = build_parser()
    args = parser.parse_args()
    init_db()
    args.func(args)


if __name__ == "__main__":
    main()
